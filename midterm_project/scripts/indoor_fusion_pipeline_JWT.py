"""
장원태 팀원 Wi‑Fi 거리 보정(RANSAC+Isotonic) · 이상치 점수(RANSAC·삼각부등식) · Huber 삼변 평가.

프로젝트 표준: `median_test_kgh_corrected`로 **학습(피팅)** 후 `median_validation`으로 **평가**한다.
단일 `--median`(구 방식): 생략 시에도 학습+검증 경로 자동 분리가 기본값이다.

원 알고리즘·파라미터는 장원태 팀원 스크립트를 따르고 입출력만 레포 규격과 맞춘다.
"""

from __future__ import annotations

import argparse
import sys
import warnings
from itertools import combinations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.optimize import least_squares
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LinearRegression, RANSACRegressor

from fusion_realtime_sanitize import (
    normalize_column_headers,
    load_sensor_tabular,
    resolve_train_kgh_corrected_paths,
    resolve_validation_paths,
)

try:
    from pandas.errors import PerformanceWarning as _JWTPerformanceWarning  # type: ignore[attr-defined]
except Exception:  # pragma: no cover
    _JWTPerformanceWarning = None  # type: ignore[misc]

if _JWTPerformanceWarning is not None:
    warnings.filterwarnings("ignore", category=_JWTPerformanceWarning)


def _fmt_dataframe_console(df: pd.DataFrame, float_decimals: int = 4) -> str:
    """콘솔 폭 안에서 숫자 열 고정 소수 출력."""
    if df is None or len(df) == 0:
        return "  (없음)"

    fmt_float = f"{{:.{float_decimals}f}}"
    out = df.copy()
    for c in out.columns:
        if pd.api.types.is_float_dtype(out[c]):
            out[c] = out[c].map(lambda x: fmt_float.format(float(x)) if pd.notna(x) else "")
        elif pd.api.types.is_integer_dtype(out[c]) and str(out[c].dtype) != "boolean":
            out[c] = out[c].map(lambda x: str(int(x)) if pd.notna(x) else "")
    return out.to_string(index=False, max_cols=24, max_colwidth=22)


def _print_block(title: str, df: pd.DataFrame) -> None:
    print(f"\n── {title} ──")
    print(_fmt_dataframe_console(df))


if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except (AttributeError, OSError):
        pass


# =========================================================
# 기본 설정 (장원태 원본 유지 — indoor_fusion_pipeline_v8.WIFI_AP_TILES 와 동일 그리드)
# =========================================================
SCALE = 0.6

wifi_devices = ["SW_11", "SW_first_team", "볼링공", "SW_4", "SW_6", "SW_5"]

anchor_positions_grid = {
    "SW_11": (1, 4),
    "SW_first_team": (20, 7),
    "볼링공": (5, 15),
    "SW_4": (4, 27),
    "SW_6": (15, 14),
    "SW_5": (14, 24),
}

anchor_positions_m = {dev: (x * SCALE, y * SCALE) for dev, (x, y) in anchor_positions_grid.items()}

wifi_anchor_xy = np.array([anchor_positions_m[dev] for dev in wifi_devices], dtype=float)

X_MIN, X_MAX = 0.0, 20 * SCALE
Y_MIN, Y_MAX = 0.0, 27 * SCALE

CORRECTION_RANSAC_THRESHOLD = 5.0
SCORE_RANSAC_THRESHOLD = 1.5

RANSAC_MAX_TRIALS = 1000
RANSAC_RANDOM_STATE = 42

BASE_TOL_DIFF = 1.0
REL_TOL_DIFF = 0.10

BASE_TOL_SUM = 0.8
REL_TOL_SUM = 0.08

TRIANGLE_NORM = 3.0
TRIANGLE_SCORE_CAP = 3.0

W_RANSAC = 3.3
W_TRIANGLE = 0.0

SCORE_WEAK = 1.5
SCORE_STRONG = 3.0
SCORE_VERY_STRONG = 5.0

MIN_WEIGHT = 0.0

HUBER_F_SCALE = 1.5

TRUE_OUTLIER_50 = 0.50
TRUE_OUTLIER_100 = 1.00


def score_label_final_outlier(s: Any) -> str:
    if pd.isna(s):
        return "계산불가"
    if s >= SCORE_VERY_STRONG:
        return "매우 강한 의심"
    if s >= SCORE_STRONG:
        return "강한 의심"
    if s >= SCORE_WEAK:
        return "약한 의심"
    return "정상"


def read_project_median(path: Path) -> pd.DataFrame:
    """레포 표준 탭 형식(CSV/XLSX) 로드 후 컬럼 정규화."""
    df = normalize_column_headers(load_sensor_tabular(Path(path)))
    unnamed_count = sum(str(c).startswith("Unnamed") for c in df.columns)
    if unnamed_count > len(df.columns) * 0.5:
        suf = Path(path).suffix.lower()
        if suf == ".csv":
            df = pd.read_csv(path, header=0, encoding="utf-8-sig")
        else:
            df = pd.read_excel(path, header=0)
        df.columns = df.columns.astype(str).str.strip()
        df = normalize_column_headers(df)
    return df


def find_col(df: pd.DataFrame, candidates):
    for c in candidates:
        if c in df.columns:
            return c
        for col in df.columns:
            if str(col).strip() == str(c).strip():
                return col
    return None


def find_device_value_col(df: pd.DataFrame, dev: str):
    candidates = [
        f"{dev}_중간값",
        f"{dev}_median",
        f"{dev}_Median",
        dev,
    ]
    return find_col(df, candidates)


def euclidean_distance(x1, y1, x2, y2):
    return np.sqrt((x1 - x2) ** 2 + (y1 - y2) ** 2)


def rmse(values):
    values = pd.Series(values).dropna()
    if len(values) == 0:
        return np.nan
    return np.sqrt(np.mean(values.astype(float) ** 2))


def mae(values):
    values = pd.Series(values).dropna()
    if len(values) == 0:
        return np.nan
    return np.mean(np.abs(values.astype(float)))


def weighted_rmse(error, weight):
    temp = pd.DataFrame({"error": error, "weight": weight}).dropna()
    temp = temp[temp["weight"] > 0]
    if len(temp) == 0:
        return np.nan
    return float(
        np.sqrt(np.sum(temp["weight"].values * temp["error"].values ** 2) / np.sum(temp["weight"].values))
    )


def make_ransac(estimator, residual_threshold):
    try:
        return RANSACRegressor(
            estimator=estimator,
            min_samples=0.5,
            residual_threshold=residual_threshold,
            max_trials=RANSAC_MAX_TRIALS,
            loss="absolute_error",
            random_state=RANSAC_RANDOM_STATE,
        )
    except TypeError:
        return RANSACRegressor(
            base_estimator=estimator,
            min_samples=0.5,
            residual_threshold=residual_threshold,
            max_trials=RANSAC_MAX_TRIALS,
            loss="absolute_loss",
            random_state=RANSAC_RANDOM_STATE,
        )


def huber_trilateration(anchor_xy, distances, weights=None):
    anchor_xy = np.asarray(anchor_xy, dtype=float)
    distances = np.asarray(distances, dtype=float)

    valid = np.isfinite(distances) & (distances > 0)

    if weights is not None:
        weights = np.asarray(weights, dtype=float)
        valid = valid & np.isfinite(weights) & (weights > 0)

    valid_count = int(valid.sum())

    if valid_count < 3:
        return np.nan, np.nan, False, np.nan, valid_count

    A = anchor_xy[valid]
    d = distances[valid]

    if weights is None:
        w = np.ones_like(d)
    else:
        w = weights[valid]

    x0 = np.average(A[:, 0], weights=w)
    y0 = np.average(A[:, 1], weights=w)
    p0 = np.array([x0, y0], dtype=float)

    def residual_func(p):
        pred = np.sqrt((p[0] - A[:, 0]) ** 2 + (p[1] - A[:, 1]) ** 2)
        residual = pred - d
        return np.sqrt(w) * residual

    try:
        res = least_squares(
            residual_func,
            p0,
            bounds=([X_MIN, Y_MIN], [X_MAX, Y_MAX]),
            loss="huber",
            f_scale=HUBER_F_SCALE,
            max_nfev=500,
        )

        return res.x[0], res.x[1], res.success, res.cost, valid_count

    except Exception:
        return np.nan, np.nan, False, np.nan, valid_count


def confusion_metrics(data: pd.DataFrame, actual_col, pred_col):
    temp = data.dropna(subset=[actual_col, pred_col]).copy()

    if len(temp) == 0:
        return {
            "전체수": 0,
            "실제이상치수": 0,
            "후보수": 0,
            "TP_잡은수": 0,
            "FP_이상치아닌데후보": 0,
            "FN_놓친수": 0,
            "TN_정상으로맞춤": 0,
            "정밀도": np.nan,
            "재현율": np.nan,
            "F1": np.nan,
        }

    actual = temp[actual_col].astype(bool)
    pred = temp[pred_col].astype(bool)

    tp = int((actual & pred).sum())
    fp = int((~actual & pred).sum())
    fn = int((actual & ~pred).sum())
    tn = int((~actual & ~pred).sum())

    precision = tp / (tp + fp) if (tp + fp) > 0 else np.nan
    recall = tp / (tp + fn) if (tp + fn) > 0 else np.nan

    f1 = (
        2 * precision * recall / (precision + recall)
        if pd.notna(precision) and pd.notna(recall) and (precision + recall) > 0
        else np.nan
    )

    return {
        "전체수": len(temp),
        "실제이상치수": int(actual.sum()),
        "후보수": int(pred.sum()),
        "TP_잡은수": tp,
        "FP_이상치아닌데후보": fp,
        "FN_놓친수": fn,
        "TN_정상으로맞춤": tn,
        "정밀도": precision,
        "재현율": recall,
        "F1": f1,
    }


def infer_column_map(train_df: pd.DataFrame) -> Tuple[str, str, Dict[str, str]]:
    node_x_col = find_col(train_df, ["Node_x", "node_x", "NODE_X", "x", "X"])
    node_y_col = find_col(train_df, ["Node_y", "node_y", "NODE_Y", "y", "Y"])
    if node_x_col is None or node_y_col is None:
        raise ValueError("train median 에서 Node_x, Node_y 컬럼을 찾지 못했습니다.")

    median_cols: Dict[str, str] = {}
    for dev in wifi_devices:
        col = find_device_value_col(train_df, dev)
        if col is None:
            raise ValueError(f"train median 에서 Wi‑Fi 컬럼 없음: {dev}")
        median_cols[dev] = col
    return node_x_col, node_y_col, median_cols


def assert_val_compatible(val_df: pd.DataFrame, node_x_col: str, node_y_col: str, median_cols: Dict[str, str]) -> None:
    if node_x_col not in val_df.columns or node_y_col not in val_df.columns:
        raise ValueError(
            "validation 과 train 에 동일한 Node 열 이름이 필요합니다 "
            f"(기대 {node_x_col}, {node_y_col}). 정규화 후 열 이름을 확인하세요."
        )
    for dev, col in median_cols.items():
        if col not in val_df.columns:
            raise ValueError(
                f"validation 에 학습 때와 같은 Wi‑Fi 열 이름이 필요합니다 ({dev} → {repr(col)})"
            )


def build_distance_frame(df: pd.DataFrame, node_x_col: str, node_y_col: str, median_cols: Dict[str, str]) -> pd.DataFrame:
    """원본 거리 · 기하 진리 거리 특성까지 채운 프레임(행 개별 인덱스 유지)."""
    result_df = df.copy()
    result_df["Node_x_m"] = pd.to_numeric(result_df[node_x_col], errors="coerce") * SCALE
    result_df["Node_y_m"] = pd.to_numeric(result_df[node_y_col], errors="coerce") * SCALE

    for dev in wifi_devices:
        result_df[f"{dev}_원본거리"] = pd.to_numeric(result_df[median_cols[dev]], errors="coerce")
        ax, ay = anchor_positions_m[dev]
        result_df[f"{dev}_실제거리"] = euclidean_distance(
            result_df["Node_x_m"],
            result_df["Node_y_m"],
            ax,
            ay,
        )
        raw = result_df[f"{dev}_원본거리"]
        true_d = result_df[f"{dev}_실제거리"]
        result_df[f"{dev}_원본_거리오차"] = raw - true_d
        result_df[f"{dev}_원본_오차율"] = np.where(true_d > 0, (raw - true_d) / true_d, np.nan)
        result_df[f"{dev}_원본_절대오차율"] = result_df[f"{dev}_원본_오차율"].abs()

    return result_df


def train_correction_iso_models(
    train_result: pd.DataFrame, *, verbose: bool = False
) -> Tuple[Dict[str, Optional[IsotonicRegression]], List[dict]]:
    """학습 데이터로 기기별 RANSAC+Isotonic 보정 피팅. train_result 에 보정 열 채움."""
    iso_by_dev: Dict[str, Optional[IsotonicRegression]] = {}
    correction_summary_rows: List[dict] = []

    for dev in wifi_devices:
        raw_col = f"{dev}_원본거리"
        true_col = f"{dev}_실제거리"

        valid = (
            train_result[raw_col].notna()
            & train_result[true_col].notna()
            & (train_result[raw_col] > 0)
            & (train_result[true_col] > 0)
        )

        train_result[f"{dev}_보정_RANSAC_Inlier"] = pd.Series(index=train_result.index, dtype="object")
        train_result[f"{dev}_보정_RANSAC_Outlier"] = pd.Series(index=train_result.index, dtype="object")
        train_result[f"{dev}_보정거리"] = np.nan
        train_result[f"{dev}_보정_거리오차"] = np.nan
        train_result[f"{dev}_보정_오차율"] = np.nan
        train_result[f"{dev}_보정_절대오차율"] = np.nan

        iso_by_dev[dev] = None

        if valid.sum() < 20:
            if verbose:
                print(f"[보정 스킵] {dev}: 학습 유효 데이터 부족")
            continue

        X = train_result.loc[valid, [raw_col]].values
        y = train_result.loc[valid, true_col].values.astype(float)
        raw_values = train_result.loc[valid, raw_col].values.astype(float)

        base_model = LinearRegression(fit_intercept=False)

        correction_ransac = make_ransac(estimator=base_model, residual_threshold=CORRECTION_RANSAC_THRESHOLD)

        correction_ransac.fit(X, y)

        inlier_mask = correction_ransac.inlier_mask_.astype(bool)
        outlier_mask = (~inlier_mask).astype(bool)

        X_inlier = raw_values[inlier_mask]
        y_inlier = y[inlier_mask]

        valid_indices = train_result.index[valid]

        train_result.loc[valid_indices, f"{dev}_보정_RANSAC_Inlier"] = inlier_mask.astype(bool).tolist()
        train_result.loc[valid_indices, f"{dev}_보정_RANSAC_Outlier"] = outlier_mask.astype(bool).tolist()

        if len(X_inlier) < 10:
            if verbose:
                print(f"[보정 스킵] {dev}: 학습 RANSAC inlier 부족 (원본거리 패스)")
            train_result.loc[valid_indices, f"{dev}_보정거리"] = raw_values
            continue

        iso_model = IsotonicRegression(increasing=True, out_of_bounds="clip")
        iso_model.fit(X_inlier, y_inlier)
        iso_by_dev[dev] = iso_model

        corrected = np.clip(iso_model.predict(raw_values), 0, None)

        train_result.loc[valid_indices, f"{dev}_보정거리"] = corrected

        true_values = y
        corr_error = corrected - true_values
        corr_error_rate = np.where(true_values > 0, corr_error / true_values, np.nan)

        train_result.loc[valid_indices, f"{dev}_보정_거리오차"] = corr_error
        train_result.loc[valid_indices, f"{dev}_보정_오차율"] = corr_error_rate
        train_result.loc[valid_indices, f"{dev}_보정_절대오차율"] = np.abs(corr_error_rate)

        raw_error = raw_values - true_values
        raw_error_rate = np.where(true_values > 0, raw_error / true_values, np.nan)

        correction_summary_rows.append(
            {
                "WiFi기기": dev,
                "유효개수": int(valid.sum()),
                "보정_RANSAC_threshold": CORRECTION_RANSAC_THRESHOLD,
                "보정_RANSAC_Inlier수": int(inlier_mask.sum()),
                "보정_RANSAC_Outlier수": int(outlier_mask.sum()),
                "보정_RANSAC_Outlier비율": float(outlier_mask.mean()),
                "원본_거리_RMSE": rmse(raw_error),
                "보정후_거리_RMSE": rmse(corr_error),
                "거리_RMSE_개선량": rmse(raw_error) - rmse(corr_error),
                "원본_평균절대오차율": np.nanmean(np.abs(raw_error_rate)),
                "보정후_평균절대오차율": np.nanmean(np.abs(corr_error_rate)),
                "원본_P90절대오차율": np.nanpercentile(np.abs(raw_error_rate), 90),
                "보정후_P90절대오차율": np.nanpercentile(np.abs(corr_error_rate), 90),
                "원본_50%이상개수": int(np.nansum(np.abs(raw_error_rate) >= 0.5)),
                "보정후_50%이상개수": int(np.nansum(np.abs(corr_error_rate) >= 0.5)),
                "원본_100%이상개수": int(np.nansum(np.abs(raw_error_rate) >= 1.0)),
                "보정후_100%이상개수": int(np.nansum(np.abs(corr_error_rate) >= 1.0)),
            }
        )

        if verbose:
            print(f"[보정 완료·학습] {dev}")
            print(f"  학습표본 원본 RMSE = {rmse(raw_error):.3f} m")
            print(f"  학습표본 보정 RMSE = {rmse(corr_error):.3f} m")
            print()

    return iso_by_dev, correction_summary_rows


def apply_correction_iso_to_eval(eval_result: pd.DataFrame, iso_by_dev: Dict[str, Optional[IsotonicRegression]]) -> None:
    """학습된 Isotonic(또는 스킵 시 원본)으로 검증 프레임에 보정 거리 채움."""
    for dev in wifi_devices:
        raw_col = f"{dev}_원본거리"
        true_col = f"{dev}_실제거리"

        eval_result[f"{dev}_보정_RANSAC_Inlier"] = pd.NA
        eval_result[f"{dev}_보정_RANSAC_Outlier"] = pd.NA
        eval_result[f"{dev}_보정거리"] = np.nan
        eval_result[f"{dev}_보정_거리오차"] = np.nan
        eval_result[f"{dev}_보정_오차율"] = np.nan
        eval_result[f"{dev}_보정_절대오차율"] = np.nan

        valid = (
            eval_result[raw_col].notna()
            & eval_result[true_col].notna()
            & (eval_result[raw_col] > 0)
            & (eval_result[true_col] > 0)
        )
        if valid.sum() == 0:
            continue

        iso = iso_by_dev.get(dev)
        raw_vals = eval_result.loc[valid, raw_col].values.astype(float)
        y_true = eval_result.loc[valid, true_col].values.astype(float)
        vi = eval_result.index[valid]

        if iso is None:
            corrected = raw_vals
        else:
            corrected = np.clip(iso.predict(raw_vals), 0, None)

        eval_result.loc[vi, f"{dev}_보정거리"] = corrected
        ce = corrected - y_true
        cer = np.where(y_true > 0, ce / y_true, np.nan)
        eval_result.loc[vi, f"{dev}_보정_거리오차"] = ce
        eval_result.loc[vi, f"{dev}_보정_오차율"] = cer
        eval_result.loc[vi, f"{dev}_보정_절대오차율"] = np.abs(cer)


def train_score_ransac_models(
    train_result: pd.DataFrame, *, verbose: bool = False
) -> Tuple[Dict[str, Optional[RANSACRegressor]], List[dict]]:
    """학습표본의 보정거리 기준 Score RANSAC 피팅."""
    score_models: Dict[str, Optional[RANSACRegressor]] = {}
    summary_rows: List[dict] = []

    for dev in wifi_devices:
        corr_col = f"{dev}_보정거리"
        true_col = f"{dev}_실제거리"

        valid = (
            train_result[corr_col].notna()
            & train_result[true_col].notna()
            & (train_result[corr_col] > 0)
            & (train_result[true_col] > 0)
        )

        score_models[dev] = None

        if valid.sum() < 20:
            if verbose:
                print(f"[score RANSAC 스킵·학습] {dev}: 유효 데이터 부족")
            continue

        X = train_result.loc[valid, [corr_col]].values
        y = train_result.loc[valid, true_col].values.astype(float)

        base_model = LinearRegression(fit_intercept=False)
        score_ransac = make_ransac(estimator=base_model, residual_threshold=SCORE_RANSAC_THRESHOLD)
        score_ransac.fit(X, y)

        pred = np.clip(score_ransac.predict(X), 0, None)
        abs_resid = np.abs(y - pred)
        inlier_mask = score_ransac.inlier_mask_.astype(bool)
        outlier_mask = (~inlier_mask).astype(bool)
        score = abs_resid / SCORE_RANSAC_THRESHOLD
        score = np.clip(score, 0, 3.0)
        score = np.where(outlier_mask, score, 0.5 * score)

        try:
            slope = float(score_ransac.estimator_.coef_[0])
        except Exception:
            slope = np.nan

        summary_rows.append(
            {
                "WiFi기기": dev,
                "유효개수": int(valid.sum()),
                "Score_RANSAC_threshold": SCORE_RANSAC_THRESHOLD,
                "Score_RANSAC_기울기": slope,
                "Score_RANSAC_Outlier수": int(outlier_mask.sum()),
                "Score_RANSAC_Outlier비율": float(outlier_mask.mean()),
                "Score_RANSAC_abs잔차_평균": float(np.mean(abs_resid)),
                "Score_RANSAC_abs잔차_P90": float(np.percentile(abs_resid, 90)),
                "RANSAC_score_평균": float(np.mean(score)),
                "RANSAC_score_P90": float(np.percentile(score, 90)),
            }
        )

        score_models[dev] = score_ransac

    return score_models, summary_rows


def apply_score_ransac_to_eval(eval_result: pd.DataFrame, score_models: Dict[str, Optional[RANSACRegressor]]) -> None:
    for dev in wifi_devices:
        corr_col = f"{dev}_보정거리"
        true_col = f"{dev}_실제거리"

        eval_result[f"{dev}_Score_RANSAC_예측실제거리"] = np.nan
        eval_result[f"{dev}_Score_RANSAC_잔차"] = np.nan
        eval_result[f"{dev}_Score_RANSAC_abs잔차"] = np.nan
        eval_result[f"{dev}_Score_RANSAC_Outlier"] = False
        eval_result[f"{dev}_RANSAC_score"] = 0.0

        mdl = score_models.get(dev)
        if mdl is None:
            continue

        valid = (
            eval_result[corr_col].notna()
            & eval_result[true_col].notna()
            & (eval_result[corr_col] > 0)
            & (eval_result[true_col] > 0)
        )
        if valid.sum() == 0:
            continue

        X = eval_result.loc[valid, [corr_col]].values
        y = eval_result.loc[valid, true_col].values.astype(float)
        pred = np.clip(mdl.predict(X), 0, None)
        abs_resid = np.abs(y - pred)
        # 검증에서는 학습 피팅 때의 inlier 마스크를 재사용할 수 없음 → 피팅 동일 residual_threshold 규준으로 라벨
        outlier_mask = abs_resid > SCORE_RANSAC_THRESHOLD

        score = abs_resid / SCORE_RANSAC_THRESHOLD
        score = np.clip(score, 0, 3.0)
        score = np.where(outlier_mask, score, 0.5 * score)

        vi = eval_result.index[valid]
        eval_result.loc[vi, f"{dev}_Score_RANSAC_예측실제거리"] = pred
        eval_result.loc[vi, f"{dev}_Score_RANSAC_잔차"] = y - pred
        eval_result.loc[vi, f"{dev}_Score_RANSAC_abs잔차"] = abs_resid
        eval_result.loc[vi, f"{dev}_Score_RANSAC_Outlier"] = outlier_mask.tolist()
        eval_result.loc[vi, f"{dev}_RANSAC_score"] = score


def finalize_evaluation_on_eval_df(result_df: pd.DataFrame, NODE_X_COL: str, NODE_Y_COL: str) -> Dict[str, Any]:
    """검증(또는 단일표) 결과에 대한 삼각부등식·종합점수·Long·위치 RMSE 산출."""
    result_df = result_df.copy()
    anchor_distances = {}

    for dev_i, dev_j in combinations(wifi_devices, 2):
        xi, yi = anchor_positions_m[dev_i]
        xj, yj = anchor_positions_m[dev_j]

        anchor_distances[(dev_i, dev_j)] = float(np.sqrt((xi - xj) ** 2 + (yi - yj) ** 2))

    def triangle_violation(d_i, d_j, L_ij):
        if pd.isna(d_i) or pd.isna(d_j):
            return 0.0, 0.0, "결측"

        if d_i <= 0 or d_j <= 0:
            return 0.0, 0.0, "비정상거리"

        tol_diff = BASE_TOL_DIFF + REL_TOL_DIFF * L_ij
        tol_sum = BASE_TOL_SUM + REL_TOL_SUM * L_ij

        diff_violation = abs(d_i - d_j) - L_ij - tol_diff
        sum_violation = L_ij - (d_i + d_j) - tol_sum

        violation_score = max(0.0, diff_violation, sum_violation)

        if violation_score <= 0:
            return 0.0, 0.0, "정상"

        if diff_violation >= sum_violation:
            ratio = violation_score / (tol_diff + 1e-9)
            return violation_score, ratio, "거리차이_위반"
        ratio = violation_score / (tol_sum + 1e-9)
        return violation_score, ratio, "거리합_위반"

    for dev in wifi_devices:
        result_df[f"{dev}_triangle_raw_score"] = 0.0
        result_df[f"{dev}_triangle_max_ratio"] = 0.0
        result_df[f"{dev}_triangle_score"] = 0.0
        result_df[f"{dev}_triangle_violation_count"] = 0

    for idx, row in result_df.iterrows():
        dev_raw_score = {dev: 0.0 for dev in wifi_devices}
        dev_max_ratio = {dev: 0.0 for dev in wifi_devices}
        dev_count = {dev: 0 for dev in wifi_devices}

        for dev_i, dev_j in combinations(wifi_devices, 2):
            d_i = row[f"{dev_i}_보정거리"]
            d_j = row[f"{dev_j}_보정거리"]
            L_ij = anchor_distances[(dev_i, dev_j)]

            v_score, v_ratio, _v_type = triangle_violation(d_i, d_j, L_ij)

            if v_score > 0:
                dev_raw_score[dev_i] += v_score
                dev_raw_score[dev_j] += v_score

                dev_max_ratio[dev_i] = max(dev_max_ratio[dev_i], v_ratio)
                dev_max_ratio[dev_j] = max(dev_max_ratio[dev_j], v_ratio)

                dev_count[dev_i] += 1
                dev_count[dev_j] += 1

        for dev in wifi_devices:
            raw_score = dev_raw_score[dev]
            max_ratio = dev_max_ratio[dev]

            tri_score = max(raw_score / TRIANGLE_NORM, max_ratio)
            tri_score = min(tri_score, TRIANGLE_SCORE_CAP)

            result_df.at[idx, f"{dev}_triangle_raw_score"] = raw_score
            result_df.at[idx, f"{dev}_triangle_max_ratio"] = max_ratio
            result_df.at[idx, f"{dev}_triangle_score"] = tri_score
            result_df.at[idx, f"{dev}_triangle_violation_count"] = dev_count[dev]

    for dev in wifi_devices:
        result_df[f"{dev}_final_outlier_score"] = W_RANSAC * result_df[f"{dev}_RANSAC_score"] + W_TRIANGLE * result_df[
            f"{dev}_triangle_score"
        ]

        result_df[f"{dev}_final_weight"] = (
            (1.0 / (1.0 + result_df[f"{dev}_final_outlier_score"])).clip(lower=MIN_WEIGHT, upper=1.0)
        )

        result_df[f"{dev}_score_약한의심"] = result_df[f"{dev}_final_outlier_score"] >= SCORE_WEAK
        result_df[f"{dev}_score_강한의심"] = result_df[f"{dev}_final_outlier_score"] >= SCORE_STRONG
        result_df[f"{dev}_score_매우강한의심"] = result_df[f"{dev}_final_outlier_score"] >= SCORE_VERY_STRONG

        result_df[f"{dev}_최종판정"] = result_df[f"{dev}_final_outlier_score"].apply(score_label_final_outlier)

    long_rows = []

    for idx, row in result_df.iterrows():
        for dev in wifi_devices:
            long_rows.append(
                {
                    "row_index": idx,
                    "Node_x": row[NODE_X_COL],
                    "Node_y": row[NODE_Y_COL],
                    "WiFi기기": dev,
                    "원본거리": row[f"{dev}_원본거리"],
                    "보정거리": row[f"{dev}_보정거리"],
                    "실제거리": row[f"{dev}_실제거리"],
                    "원본_거리오차": row[f"{dev}_원본_거리오차"],
                    "보정_거리오차": row[f"{dev}_보정_거리오차"],
                    "원본_오차율": row[f"{dev}_원본_오차율"],
                    "보정_오차율": row[f"{dev}_보정_오차율"],
                    "원본_절대오차율": row[f"{dev}_원본_절대오차율"],
                    "보정_절대오차율": row[f"{dev}_보정_절대오차율"],
                    "RANSAC_score": row[f"{dev}_RANSAC_score"],
                    "triangle_score": row[f"{dev}_triangle_score"],
                    "final_outlier_score": row[f"{dev}_final_outlier_score"],
                    "final_weight": row[f"{dev}_final_weight"],
                    "최종판정": row[f"{dev}_최종판정"],
                    "score_약한의심": row[f"{dev}_score_약한의심"],
                    "score_강한의심": row[f"{dev}_score_강한의심"],
                    "score_매우강한의심": row[f"{dev}_score_매우강한의심"],
                    "원본_실제이상치_50": row[f"{dev}_원본_절대오차율"] >= TRUE_OUTLIER_50
                    if pd.notna(row[f"{dev}_원본_절대오차율"])
                    else np.nan,
                    "보정_실제이상치_50": row[f"{dev}_보정_절대오차율"] >= TRUE_OUTLIER_50
                    if pd.notna(row[f"{dev}_보정_절대오차율"])
                    else np.nan,
                    "원본_실제이상치_100": row[f"{dev}_원본_절대오차율"] >= TRUE_OUTLIER_100
                    if pd.notna(row[f"{dev}_원본_절대오차율"])
                    else np.nan,
                    "보정_실제이상치_100": row[f"{dev}_보정_절대오차율"] >= TRUE_OUTLIER_100
                    if pd.notna(row[f"{dev}_보정_절대오차율"])
                    else np.nan,
                }
            )

    long_df = pd.DataFrame(long_rows)

    score_eval_rows = []

    for pred_name, pred_col in [
        ("score>=약한의심", "score_약한의심"),
        ("score>=강한의심", "score_강한의심"),
        ("score>=매우강한의심", "score_매우강한의심"),
    ]:
        for actual_name, actual_col in [
            ("원본_절대오차율_50이상", "원본_실제이상치_50"),
            ("보정_절대오차율_50이상", "보정_실제이상치_50"),
            ("원본_절대오차율_100이상", "원본_실제이상치_100"),
            ("보정_절대오차율_100이상", "보정_실제이상치_100"),
        ]:
            m = confusion_metrics(long_df, actual_col, pred_col)
            m["후보기준"] = pred_name
            m["실제기준"] = actual_name
            score_eval_rows.append(m)

    score_eval_df = pd.DataFrame(score_eval_rows)
    score_eval_df = score_eval_df[
        [
            "후보기준",
            "실제기준",
            "전체수",
            "실제이상치수",
            "후보수",
            "TP_잡은수",
            "FP_이상치아닌데후보",
            "FN_놓친수",
            "TN_정상으로맞춤",
            "정밀도",
            "재현율",
            "F1",
        ]
    ]

    distance_rmse_rows = []

    valid_all = long_df.dropna(subset=["원본_거리오차", "보정_거리오차"]).copy()
    hard_kept = valid_all[valid_all["score_강한의심"] == False].copy()

    distance_rmse_rows.append(
        {
            "구분": "전체",
            "유효개수": len(valid_all),
            "원본_거리_RMSE": rmse(valid_all["원본_거리오차"]),
            "보정후_거리_RMSE": rmse(valid_all["보정_거리오차"]),
            "보정후_soft_weighted_거리_RMSE": weighted_rmse(valid_all["보정_거리오차"], valid_all["final_weight"]),
            "보정후_hard제거_거리_RMSE": rmse(hard_kept["보정_거리오차"]),
            "원본_거리_MAE": mae(valid_all["원본_거리오차"]),
            "보정후_거리_MAE": mae(valid_all["보정_거리오차"]),
            "원본_평균절대오차율": valid_all["원본_절대오차율"].mean(),
            "보정후_평균절대오차율": valid_all["보정_절대오차율"].mean(),
            "보정후_soft_평균weight": valid_all["final_weight"].mean(),
        }
    )

    for dev in wifi_devices:
        sub = valid_all[valid_all["WiFi기기"] == dev].copy()
        sub_hard_kept = sub[sub["score_강한의심"] == False].copy()

        distance_rmse_rows.append(
            {
                "구분": dev,
                "유효개수": len(sub),
                "원본_거리_RMSE": rmse(sub["원본_거리오차"]),
                "보정후_거리_RMSE": rmse(sub["보정_거리오차"]),
                "보정후_soft_weighted_거리_RMSE": weighted_rmse(sub["보정_거리오차"], sub["final_weight"]),
                "보정후_hard제거_거리_RMSE": rmse(sub_hard_kept["보정_거리오차"]),
                "원본_거리_MAE": mae(sub["원본_거리오차"]),
                "보정후_거리_MAE": mae(sub["보정_거리오차"]),
                "원본_평균절대오차율": sub["원본_절대오차율"].mean(),
                "보정후_평균절대오차율": sub["보정_절대오차율"].mean(),
                "보정후_soft_평균weight": sub["final_weight"].mean(),
            }
        )

    distance_rmse_df = pd.DataFrame(distance_rmse_rows)

    result_df["Raw_Huber_x"] = np.nan
    result_df["Raw_Huber_y"] = np.nan
    result_df["Raw_Huber_success"] = False
    result_df["Raw_Huber_위치오차"] = np.nan

    result_df["Corrected_Huber_x"] = np.nan
    result_df["Corrected_Huber_y"] = np.nan
    result_df["Corrected_Huber_success"] = False
    result_df["Corrected_Huber_위치오차"] = np.nan

    result_df["Corrected_Soft_x"] = np.nan
    result_df["Corrected_Soft_y"] = np.nan
    result_df["Corrected_Soft_success"] = False
    result_df["Corrected_Soft_위치오차"] = np.nan

    result_df["Corrected_Hard_x"] = np.nan
    result_df["Corrected_Hard_y"] = np.nan
    result_df["Corrected_Hard_success"] = False
    result_df["Corrected_Hard_위치오차"] = np.nan
    result_df["Corrected_Hard_사용기기수"] = 0

    for idx, row in result_df.iterrows():
        raw_distances = np.array([row[f"{dev}_원본거리"] for dev in wifi_devices], dtype=float)

        corrected_distances = np.array([row[f"{dev}_보정거리"] for dev in wifi_devices], dtype=float)

        soft_weights = np.array([row[f"{dev}_final_weight"] for dev in wifi_devices], dtype=float)

        true_x = row["Node_x_m"]
        true_y = row["Node_y_m"]

        xr, yr, success_r, _cost_r, _valid_r = huber_trilateration(wifi_anchor_xy, raw_distances, weights=None)

        result_df.at[idx, "Raw_Huber_x"] = xr
        result_df.at[idx, "Raw_Huber_y"] = yr
        result_df.at[idx, "Raw_Huber_success"] = success_r

        if success_r and np.isfinite(xr) and np.isfinite(yr):
            result_df.at[idx, "Raw_Huber_위치오차"] = float(np.sqrt((xr - true_x) ** 2 + (yr - true_y) ** 2))

        xc, yc, success_c, _cost_c, _valid_c = huber_trilateration(wifi_anchor_xy, corrected_distances, weights=None)

        result_df.at[idx, "Corrected_Huber_x"] = xc
        result_df.at[idx, "Corrected_Huber_y"] = yc
        result_df.at[idx, "Corrected_Huber_success"] = success_c

        if success_c and np.isfinite(xc) and np.isfinite(yc):
            result_df.at[idx, "Corrected_Huber_위치오차"] = float(np.sqrt((xc - true_x) ** 2 + (yc - true_y) ** 2))

        xs, ys, success_s, _cost_s, _valid_s = huber_trilateration(wifi_anchor_xy, corrected_distances, weights=soft_weights)

        result_df.at[idx, "Corrected_Soft_x"] = xs
        result_df.at[idx, "Corrected_Soft_y"] = ys
        result_df.at[idx, "Corrected_Soft_success"] = success_s

        if success_s and np.isfinite(xs) and np.isfinite(ys):
            result_df.at[idx, "Corrected_Soft_위치오차"] = float(np.sqrt((xs - true_x) ** 2 + (ys - true_y) ** 2))

        hard_keep = np.array([not bool(row[f"{dev}_score_강한의심"]) for dev in wifi_devices], dtype=bool)

        valid_distance = np.isfinite(corrected_distances) & (corrected_distances > 0)
        hard_keep = hard_keep & valid_distance

        if hard_keep.sum() < 3:
            hard_weights = valid_distance.astype(float)
        else:
            hard_weights = hard_keep.astype(float)

        xh, yh, success_h, _cost_h, _valid_h = huber_trilateration(wifi_anchor_xy, corrected_distances, weights=hard_weights)

        result_df.at[idx, "Corrected_Hard_x"] = xh
        result_df.at[idx, "Corrected_Hard_y"] = yh
        result_df.at[idx, "Corrected_Hard_success"] = success_h
        result_df.at[idx, "Corrected_Hard_사용기기수"] = int((hard_weights > 0).sum())

        if success_h and np.isfinite(xh) and np.isfinite(yh):
            result_df.at[idx, "Corrected_Hard_위치오차"] = float(np.sqrt((xh - true_x) ** 2 + (yh - true_y) ** 2))

    position_rmse_rows = []

    for method_name, err_col in [
        ("원본거리_Huber", "Raw_Huber_위치오차"),
        ("보정거리_Huber", "Corrected_Huber_위치오차"),
        ("보정거리_soft_weight", "Corrected_Soft_위치오차"),
        ("보정거리_hard_remove", "Corrected_Hard_위치오차"),
    ]:
        e = result_df[err_col]

        position_rmse_rows.append(
            {
                "방식": method_name,
                "유효행수": int(e.notna().sum()),
                "위치_RMSE": rmse(e),
                "위치_MAE": float(e.mean()) if e.notna().any() else np.nan,
                "위치_P90": float(e.quantile(0.9)) if e.notna().sum() >= 2 else np.nan,
                "위치_최대오차": float(e.max()) if e.notna().any() else np.nan,
            }
        )

    position_rmse_df = pd.DataFrame(position_rmse_rows)

    same_valid = result_df.dropna(
        subset=[
            "Raw_Huber_위치오차",
            "Corrected_Huber_위치오차",
            "Corrected_Soft_위치오차",
            "Corrected_Hard_위치오차",
        ]
    ).copy()

    same_position_rows = []

    if len(same_valid) > 0:
        raw_rmse_row = rmse(same_valid["Raw_Huber_위치오차"])

        same_position_rows.append(
            {
                "비교기준": "동일행",
                "행수": len(same_valid),
                "원본_RMSE": raw_rmse_row,
                "보정_RMSE": rmse(same_valid["Corrected_Huber_위치오차"]),
                "보정_soft_RMSE": rmse(same_valid["Corrected_Soft_위치오차"]),
                "보정_hard_RMSE": rmse(same_valid["Corrected_Hard_위치오차"]),
                "보정_RMSE감소량": raw_rmse_row - rmse(same_valid["Corrected_Huber_위치오차"]),
                "보정_soft_RMSE감소량": raw_rmse_row - rmse(same_valid["Corrected_Soft_위치오차"]),
                "보정_hard_RMSE감소량": raw_rmse_row - rmse(same_valid["Corrected_Hard_위치오차"]),
                "원본_MAE": float(same_valid["Raw_Huber_위치오차"].mean()),
                "보정_MAE": float(same_valid["Corrected_Huber_위치오차"].mean()),
                "보정_soft_MAE": float(same_valid["Corrected_Soft_위치오차"].mean()),
                "원본_P90": float(same_valid["Raw_Huber_위치오차"].quantile(0.9)),
                "보정_P90": float(same_valid["Corrected_Huber_위치오차"].quantile(0.9)),
                "보정_soft_P90": float(same_valid["Corrected_Soft_위치오차"].quantile(0.9)),
            }
        )

    same_position_df = pd.DataFrame(same_position_rows)

    return {
        "result_df": result_df,
        "long_df": long_df,
        "score_eval_df": score_eval_df,
        "distance_rmse_df": distance_rmse_df,
        "position_rmse_df": position_rmse_df,
        "same_position_df": same_position_df,
    }


def run_train_val_analysis(
    train_median_df: pd.DataFrame,
    val_median_df: pd.DataFrame,
    *,
    train_median_path: Optional[Path] = None,
    val_median_path: Optional[Path] = None,
    verbose: bool = False,
) -> Dict[str, Any]:
    """학습 median으로 보정·Score RANSAC 피팅 후, 검증 median으로 평가."""
    NODE_X_COL, NODE_Y_COL, median_cols = infer_column_map(train_median_df)
    assert_val_compatible(val_median_df, NODE_X_COL, NODE_Y_COL, median_cols)

    if verbose:
        print("[학습] Node 열:", NODE_X_COL, NODE_Y_COL)
        print("[학습] Wi‑Fi 열:", median_cols)

    train_result = build_distance_frame(train_median_df, NODE_X_COL, NODE_Y_COL, median_cols)
    iso_by_dev, corr_rows = train_correction_iso_models(train_result, verbose=verbose)
    correction_summary_df = pd.DataFrame(corr_rows)

    val_result = build_distance_frame(val_median_df, NODE_X_COL, NODE_Y_COL, median_cols)
    apply_correction_iso_to_eval(val_result, iso_by_dev)

    score_models, score_rows = train_score_ransac_models(train_result, verbose=verbose)
    ransac_score_summary_df = pd.DataFrame(score_rows)
    apply_score_ransac_to_eval(val_result, score_models)

    out = finalize_evaluation_on_eval_df(val_result, NODE_X_COL, NODE_Y_COL)
    out["correction_summary_df"] = correction_summary_df
    out["ransac_score_summary_df"] = ransac_score_summary_df

    base_param = [
        {"항목": "CORRECTION_RANSAC_THRESHOLD", "값": CORRECTION_RANSAC_THRESHOLD},
        {"항목": "SCORE_RANSAC_THRESHOLD", "값": SCORE_RANSAC_THRESHOLD},
        {"항목": "BASE_TOL_DIFF", "값": BASE_TOL_DIFF},
        {"항목": "REL_TOL_DIFF", "값": REL_TOL_DIFF},
        {"항목": "BASE_TOL_SUM", "값": BASE_TOL_SUM},
        {"항목": "REL_TOL_SUM", "값": REL_TOL_SUM},
        {"항목": "HUBER_F_SCALE", "값": HUBER_F_SCALE},
        {"항목": "W_RANSAC", "값": W_RANSAC},
        {"항목": "W_TRIANGLE", "값": W_TRIANGLE},
        {"항목": "SCORE_WEAK", "값": SCORE_WEAK},
        {"항목": "SCORE_STRONG", "값": SCORE_STRONG},
        {"항목": "SCORE_VERY_STRONG", "값": SCORE_VERY_STRONG},
        {"항목": "MIN_WEIGHT", "값": MIN_WEIGHT},
        {
            "항목": "학습피팅_행수(train)",
            "값": len(train_median_df),
        },
        {
            "항목": "평가_행수(validation)",
            "값": len(val_median_df),
        },
        {
            "항목": "행별·Long·거리_RMSE·위치_RMSE·점수탐지_시트",
            "값": "검증(validation median) 결과",
        },
    ]
    if train_median_path is not None:
        base_param.append({"항목": "train_median_path", "값": str(train_median_path)})
    if val_median_path is not None:
        base_param.append({"항목": "val_median_path", "값": str(val_median_path)})

    out["param_df"] = pd.DataFrame(base_param)

    out["train_fit_rowwise_df"] = train_result.copy()

    return out


def save_styled_excel(output_file: Path, tables: Dict[str, Any]) -> None:
    result_df = tables["result_df"]
    long_df = tables["long_df"]
    correction_summary_df = tables["correction_summary_df"]
    ransac_score_summary_df = tables["ransac_score_summary_df"]
    score_eval_df = tables["score_eval_df"]
    distance_rmse_df = tables["distance_rmse_df"]
    position_rmse_df = tables["position_rmse_df"]
    same_position_df = tables["same_position_df"]
    param_df = tables["param_df"]

    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="행별_보정_점수결과", index=False)
        long_df.to_excel(writer, sheet_name="기기별_Long결과", index=False)

        correction_summary_df.to_excel(writer, sheet_name="보정모델_요약", index=False)
        ransac_score_summary_df.to_excel(writer, sheet_name="Score_RANSAC_요약", index=False)

        score_eval_df.to_excel(writer, sheet_name="점수_탐지성능", index=False)
        distance_rmse_df.to_excel(writer, sheet_name="거리_RMSE_비교", index=False)
        position_rmse_df.to_excel(writer, sheet_name="위치_RMSE_비교", index=False)
        same_position_df.to_excel(writer, sheet_name="동일행_위치비교", index=False)
        param_df.to_excel(writer, sheet_name="파라미터", index=False)

        train_fit = tables.get("train_fit_rowwise_df")
        if train_fit is not None and len(train_fit) > 0:
            train_fit.to_excel(writer, sheet_name="학습_fit_행별", index=False)

        wb = writer.book

        thin_border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        )

        header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        good_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
        weak_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        strong_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
        very_strong_fill = PatternFill(fill_type="solid", fgColor="E06666")

        def style_sheet(ws):
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                    if isinstance(cell.value, float):
                        cell.number_format = "0.000"

            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_length = 0
                col_letter = get_column_letter(col_idx)

                for cell in col_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = min(max_length + 2, 36)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        for sheet_name in wb.sheetnames:
            style_sheet(wb[sheet_name])

        for sheet_name in ["행별_보정_점수결과", "기기별_Long결과"]:
            ws = wb[sheet_name]

            for col_idx in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col_idx).value)

                if header.endswith("_최종판정") or header == "최종판정":
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)

                        if cell.value == "정상":
                            cell.fill = good_fill
                        elif cell.value == "약한 의심":
                            cell.fill = weak_fill
                        elif cell.value == "강한 의심":
                            cell.fill = strong_fill
                            cell.font = Font(bold=True, color="C00000")
                        elif cell.value == "매우 강한 의심":
                            cell.fill = very_strong_fill
                            cell.font = Font(bold=True, color="FFFFFF")

                if "score_강한의심" in header or "score_매우강한의심" in header:
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is True:
                            cell.fill = strong_fill
                            cell.font = Font(bold=True, color="C00000")


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    train_dir = root / "data" / "train"
    val_dir = root / "data" / "validation"
    default_out = root / "outputs" / "jwt_wifi_corrected_outlier_score_position_rmse.xlsx"

    parser = argparse.ArgumentParser(description="장원태 팀원 Wi‑Fi 파이프라인 (train 피팅 → validation 평가)")
    parser.add_argument(
        "--train-median",
        type=Path,
        default=None,
        help="학습용 median CSV/XLSX (기본: median_test_kgh_corrected)",
    )
    parser.add_argument(
        "--val-median",
        type=Path,
        default=None,
        help="평가용 median CSV/XLSX (기본: median_validation)",
    )
    parser.add_argument("--median", type=Path, default=None, help="단일표만 처리(학습+평가 동일 데이터, 비추전)")
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="저장 xlsx (기본: outputs/jwt_wifi_corrected_outlier_score_position_rmse.xlsx)",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="기기별 학습 진행 로그·전체 표 출력 (기본은 요약만)",
    )
    args = parser.parse_args()

    output_path = Path(args.output).resolve() if args.output is not None else default_out.resolve()

    if args.median is not None:
        single = Path(args.median).resolve()
        print("\n[JWT] --median 단일표: 학습·평가가 동일 표 (일반 검증은 옵션 없이 실행)")
        print(f"  파일: {single}")
        median_df = read_project_median(single)
        tables = run_train_val_analysis(
            median_df,
            median_df.copy(),
            train_median_path=single,
            val_median_path=single,
            verbose=args.verbose,
        )
    else:
        train_path = Path(args.train_median).resolve() if args.train_median else resolve_train_kgh_corrected_paths(train_dir)[0]
        val_path = Path(args.val_median).resolve() if args.val_median else resolve_validation_paths(val_dir)[0]
        train_df = read_project_median(train_path)
        val_df = read_project_median(val_path)
        print("\n[JWT] train 피팅 → validation 평가")
        print(f"  학습: {train_path}")
        print(f"  검증: {val_path}")
        print(f"  저장: {output_path}")
        tables = run_train_val_analysis(
            train_df,
            val_df,
            train_median_path=train_path,
            val_median_path=val_path,
            verbose=args.verbose,
        )

    save_styled_excel(output_path, tables)

    cdf = tables["correction_summary_df"]
    sdf = tables["ransac_score_summary_df"]

    if not args.verbose and len(cdf) < len(wifi_devices):
        done = set(cdf["WiFi기기"].astype(str).tolist()) if len(cdf) and "WiFi기기" in cdf.columns else set()
        pend = [d for d in wifi_devices if d not in done]
        if pend:
            print(f"\n[학습] 보정 피팅 미수행 기기: {', '.join(pend)}")

    print(f"\n완료 — 엑셀 저장: {output_path}")

    if args.verbose:
        _print_block("학습 | 보정 요약 (전체 열)", cdf)
        _print_block("학습 | Score RANSAC 피팅 요약 (전체 열)", sdf)
        _print_block("검증 | 거리 RMSE 비교", tables["distance_rmse_df"])
        _print_block("검증 | 위치 RMSE 비교", tables["position_rmse_df"])
        _print_block("검증 | 동일행 위치 비교", tables["same_position_df"])
    else:
        corr_keeps = ["WiFi기기", "유효개수", "원본_거리_RMSE", "보정후_거리_RMSE", "거리_RMSE_개선량"]
        corr_sub = cdf[[c for c in corr_keeps if c in cdf.columns]]
        score_keeps = ["WiFi기기", "유효개수", "Score_RANSAC_기울기", "RANSAC_score_평균", "RANSAC_score_P90"]
        score_sub = sdf[[c for c in score_keeps if c in sdf.columns]]
        _print_block("학습 | 거리 보정 (in-sample, 핵심 열만 — 전체는 엑셀·--verbose)", corr_sub)
        _print_block("학습 | Score RANSAC (핵심 열만)", score_sub)
        _print_block("검증 | 거리 RMSE 비교", tables["distance_rmse_df"])
        _print_block("검증 | 위치 RMSE 비교", tables["position_rmse_df"])
        _print_block("검증 | 동일행 위치 비교", tables["same_position_df"])

    if not args.verbose:
        print("(--verbose 로 학습 과정 로그·전열 표 출력)")

    return 0


if __name__ == "__main__":
    from pathlib import Path

    from script_run_io import cli_entrypoint

    cli_entrypoint(Path(__file__), main, output_artifact_include_prefixes=("jwt_",))
